"""
Sales IQ - Data Import Service
Handles CSV and Excel (XLSX/XLS) upload, field mapping, validation, preview,
and bulk import with automatic Data Quality Agent integration.
"""

import csv
import io
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID, uuid4

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.core import AuditLog
from app.models.business import (
    Customer, Invoice, Payment, Dispute, CollectionActivity, CreditLimitRequest,
    Product,
    CustomerStatus, InvoiceStatus, PaymentMethod,
    DisputeStatus, DisputeReason, CollectionAction, CreditApprovalStatus,
    DataQualityRecord, DataQualityStatus,
)


# =============================================
# Field Definitions per Entity
# =============================================

ENTITY_FIELDS = {
    "customers": {
        # field_name: (db_column, type, required, description)
        "name": ("name", "str", True, "Company name"),
        "name_ar": ("name_ar", "str", False, "Arabic company name"),
        "trade_name": ("trade_name", "str", False, "Trade / DBA name"),
        "external_id": ("external_id", "str", False, "External system ID"),
        "tax_id": ("tax_id", "str", False, "Tax registration number (TRN)"),
        "status": ("status", "str", False, "active / inactive / blocked"),
        "industry": ("industry", "str", False, "Industry sector"),
        "segment": ("segment", "str", False, "Customer segment"),
        "territory": ("territory", "str", False, "Sales territory"),
        "region": ("region", "str", False, "Region"),
        "country": ("country", "str", False, "ISO country code (AE, SA, ...)"),
        "city": ("city", "str", False, "City"),
        "address": ("address", "str", False, "Full address"),
        "phone": ("phone", "str", False, "Phone number"),
        "email": ("email", "str", False, "Email address"),
        "website": ("website", "str", False, "Website URL"),
        "currency": ("currency", "str", False, "Default currency (AED, SAR, ...)"),
        "payment_terms_days": ("payment_terms_days", "int", False, "Payment terms in days"),
        "credit_limit": ("credit_limit", "decimal", False, "Credit limit amount"),
    },
    "invoices": {
        "customer_external_id": (None, "str", True, "Customer external ID (for lookup)"),
        "customer_name": (None, "str", False, "Customer name (fallback lookup)"),
        "invoice_number": ("invoice_number", "str", True, "Invoice number"),
        "external_id": ("external_id", "str", False, "External system ID"),
        "po_number": ("po_number", "str", False, "Purchase order number"),
        "invoice_date": ("invoice_date", "date", True, "Invoice date (YYYY-MM-DD)"),
        "due_date": ("due_date", "date", True, "Payment due date (YYYY-MM-DD)"),
        "currency": ("currency", "str", False, "Currency code"),
        "amount": ("amount", "decimal", True, "Invoice amount"),
        "tax_amount": ("tax_amount", "decimal", False, "Tax amount"),
        "discount_amount": ("discount_amount", "decimal", False, "Discount amount"),
        "status": ("status", "str", False, "open / paid / overdue / ..."),
        "notes": ("notes", "str", False, "Notes"),
    },
    "payments": {
        "customer_external_id": (None, "str", True, "Customer external ID (for lookup)"),
        "customer_name": (None, "str", False, "Customer name (fallback lookup)"),
        "invoice_number": (None, "str", False, "Invoice number (for matching)"),
        "external_id": ("external_id", "str", False, "External system ID"),
        "payment_date": ("payment_date", "date", True, "Payment date (YYYY-MM-DD)"),
        "amount": ("amount", "decimal", True, "Payment amount"),
        "currency": ("currency", "str", False, "Currency code"),
        "payment_method": ("payment_method", "str", False, "bank_transfer / check / cash / ..."),
        "reference_number": ("reference_number", "str", False, "Payment reference"),
        "bank_reference": ("bank_reference", "str", False, "Bank reference number"),
        "notes": ("notes", "str", False, "Notes"),
    },
    "collections": {
        "customer_external_id": (None, "str", True, "Customer external ID (for lookup)"),
        "customer_name": (None, "str", False, "Customer name (fallback lookup)"),
        "invoice_number": (None, "str", False, "Invoice number (for matching)"),
        "action_type": ("action_type", "str", True, "email_reminder / phone_call / escalation / ..."),
        "action_date": ("action_date", "date", True, "Action date (YYYY-MM-DD)"),
        "notes": ("notes", "str", False, "Notes"),
        "ptp_date": ("ptp_date", "date", False, "Promise-to-pay date"),
        "ptp_amount": ("ptp_amount", "decimal", False, "Promise-to-pay amount"),
    },
    "disputes": {
        "customer_external_id": (None, "str", True, "Customer external ID (for lookup)"),
        "customer_name": (None, "str", False, "Customer name (fallback lookup)"),
        "invoice_number": (None, "str", False, "Invoice number (for matching)"),
        "dispute_number": ("dispute_number", "str", True, "Dispute reference number"),
        "reason": ("reason", "str", True, "pricing / quantity / quality / delivery / duplicate / ..."),
        "reason_detail": ("reason_detail", "str", False, "Detailed reason description"),
        "amount": ("amount", "decimal", True, "Dispute amount"),
        "currency": ("currency", "str", False, "Currency code"),
        "status": ("status", "str", False, "open / in_review / resolved / ..."),
        "priority": ("priority", "str", False, "low / medium / high / critical"),
        "sla_due_date": ("sla_due_date", "date", False, "SLA due date"),
    },
    "credit_limits": {
        "customer_external_id": (None, "str", True, "Customer external ID (for lookup)"),
        "customer_name": (None, "str", False, "Customer name (fallback lookup)"),
        "requested_limit": ("requested_limit", "decimal", True, "Requested credit limit"),
        "currency": ("currency", "str", False, "Currency code"),
        "justification": ("justification", "str", False, "Request justification"),
        "approval_status": ("approval_status", "str", False, "pending / approved / rejected"),
    },
}

# Common header aliases for auto-mapping
HEADER_ALIASES = {
    # Customer aliases
    "company": "name", "company_name": "name", "customer_name": "name",
    "customer": "name", "account_name": "name",
    "arabic_name": "name_ar", "name_arabic": "name_ar",
    "trn": "tax_id", "vat_number": "tax_id", "tax_number": "tax_id",
    "sector": "industry", "type": "segment", "customer_type": "segment",
    "area": "territory", "zone": "territory", "sales_territory": "territory",
    "tel": "phone", "telephone": "phone", "mobile": "phone", "contact_number": "phone",
    "mail": "email", "email_address": "email", "contact_email": "email",
    "terms": "payment_terms_days", "payment_terms": "payment_terms_days",
    "credit": "credit_limit", "limit": "credit_limit",
    "ext_id": "external_id", "erp_id": "external_id", "customer_id": "external_id",
    "account_id": "external_id", "code": "external_id", "customer_code": "external_id",

    # Invoice aliases
    "inv_number": "invoice_number", "inv_no": "invoice_number", "invoice_no": "invoice_number",
    "inv_date": "invoice_date", "date": "invoice_date", "doc_date": "invoice_date",
    "due": "due_date", "pay_by": "due_date", "maturity_date": "due_date",
    "total": "amount", "total_amount": "amount", "net_amount": "amount",
    "inv_amount": "amount", "gross_amount": "amount",
    "tax": "tax_amount", "vat": "tax_amount", "vat_amount": "tax_amount",
    "discount": "discount_amount",
    "po": "po_number", "purchase_order": "po_number",
    "cust_id": "customer_external_id", "cust_code": "customer_external_id",
    "account_code": "customer_external_id",

    # Payment aliases
    "pay_date": "payment_date", "receipt_date": "payment_date", "value_date": "payment_date",
    "pay_amount": "amount", "received": "amount", "receipt_amount": "amount",
    "method": "payment_method", "pay_method": "payment_method", "mode": "payment_method",
    "ref": "reference_number", "payment_ref": "reference_number",
    "bank_ref": "bank_reference", "transaction_ref": "bank_reference",
    "inv_no": "invoice_number", "against_invoice": "invoice_number",

    # Collection aliases
    "action": "action_type", "collection_type": "action_type", "activity_type": "action_type",
    "collection_action": "action_type", "call_type": "action_type",
    "collection_date": "action_date", "activity_date": "action_date",
    "promise_date": "ptp_date", "ptp": "ptp_date", "promise_to_pay_date": "ptp_date",
    "promise_amount": "ptp_amount", "ptp_value": "ptp_amount",

    # Dispute aliases
    "dispute_no": "dispute_number", "disp_number": "dispute_number", "disp_no": "dispute_number",
    "case_number": "dispute_number", "dispute_ref": "dispute_number",
    "dispute_reason": "reason", "disp_reason": "reason",
    "dispute_amount": "amount", "disp_amount": "amount",
    "dispute_status": "status", "disp_status": "status",
    "sla_date": "sla_due_date", "resolution_due": "sla_due_date",

    # Credit limit aliases
    "new_limit": "requested_limit", "req_limit": "requested_limit", "credit_request": "requested_limit",
    "new_credit_limit": "requested_limit", "proposed_limit": "requested_limit",
    "reason": "justification", "request_reason": "justification",
    "status": "approval_status",

    # ── D365 Finance & Operations ──────────────────────────────────
    # Customer Master
    "account": "external_id",
    "account_num": "external_id",
    "customer_account": "external_id",
    "cust_account": "external_id",
    "invoice_account": "external_id",
    "customer_group": "industry",
    "customer_group_name": "industry",
    "customer_classification": "segment",
    "price_group": "segment",
    "warehouse": "territory",
    "default_warehouse": "territory",
    "site": "territory",
    "sales_district": "territory",
    "terms_of_payment": "payment_terms_days",
    "payment_terms_name": "payment_terms_days",
    "payment_term": "payment_terms_days",
    "payment_condition": "payment_terms_days",
    "account_status": "status",
    "customer_hold": "status",
    "tax_exempt_number": "tax_id",
    "tax_registration_number": "tax_id",
    "vat_registration_number": "tax_id",
    "registration_number": "tax_id",
    "sales_currency": "currency",
    "default_currency": "currency",
    "currency_code": "currency",
    "credit_max": "credit_limit",
    "credit_limit_amount": "credit_limit",
    "delivery_address": "address",
    "primary_address": "address",
    "street": "address",
    "line_of_business": "industry",
    "organization_name": "name",
    "legal_entity_name": "name",
    "party_name": "name",
    "debtor": "name",
    "customer_name_ar": "name_ar",
    "name_in_arabic": "name_ar",
    "arabic_company_name": "name_ar",
    "web_address": "website",
    "internet_address": "website",
    "region_code": "region",
    "country_code": "country",
    "country_region": "country",
    "state": "region",
    "province": "region",
    "employee_responsible": "segment",
    "sales_responsible": "segment",
    "sales_rep": "segment",
    "delivery_terms": "_skip",
    "method_of_payment": "_skip",
    "payment_mode": "_skip",
    "sales_tax_group": "_skip",
    "account_number": "_skip",
    "created_by": "_skip",
    "created_date_and_time": "_skip",
    "create_intercompany_orders": "_skip",
    "exclude_from_credit_management": "_skip",
    "mandatory_credit_limit": "_skip",

    # D365 Sales Order / Invoice
    "sales_order": "invoice_number",
    "order_number": "invoice_number",
    "sales_id": "invoice_number",
    "so_number": "invoice_number",
    "document_number": "invoice_number",
    "voucher": "invoice_number",
    "document_date": "invoice_date",
    "posting_date": "invoice_date",
    "transaction_date": "invoice_date",
    "delivery_date": "due_date",
    "confirmed_ship_date": "due_date",
    "requested_ship_date": "due_date",
    "line_amount": "amount",
    "line_net_amount": "amount",
    "sales_amount": "amount",
    "order_amount": "amount",
    "original_amount": "amount",
    "amount_in_transaction_currency": "amount",
    "line_status": "status",
    "order_status": "status",
    "delivery_name": "customer_name",
    "invoice_customer": "customer_external_id",
    "order_account": "customer_external_id",
    "invoicing_customer": "customer_external_id",
    "customer_reference": "po_number",

    # D365 Payment / AR
    "settled_date": "payment_date",
    "settlement_date": "payment_date",
    "last_settlement_date": "payment_date",
    "clearing_date": "payment_date",
    "amount_settled": "amount",
    "settlement_amount": "amount",
    "payment_reference": "reference_number",
    "journal_number": "reference_number",
    "payment_journal": "reference_number",
    "bank_transaction_id": "bank_reference",
    "payment_type": "payment_method",
    "method_of_payment": "payment_method",
    "payment_specification": "payment_method",

    # ── SAP Business One ───────────────────────────────────────────
    "bp_code": "external_id",
    "card_code": "external_id",
    "card_name": "name",
    "group_name": "industry",
    "doc_entry": "external_id",
    "doc_num": "invoice_number",
    "doc_date": "invoice_date",
    "doc_due_date": "due_date",
    "doc_total": "amount",
    "card_foreign_name": "name_ar",
    "debtor_code": "external_id",
    "debtor_name": "name",
}


# =============================================
# Excel Parser (XLSX / XLS)
# =============================================

class ExcelParser:
    """Parse Excel files (.xlsx, .xls) using openpyxl with smart header detection."""

    @staticmethod
    def _detect_header_row(raw_rows: list, max_scan: int = 20) -> int:
        """
        Detect the actual header row in Excel data.
        D365 report exports often have title / subtitle / blank rows before data.
        The header row is the one with the most text (non-numeric) cells.
        """
        best_idx = 0
        best_score = 0

        for i, row in enumerate(raw_rows[:max_scan]):
            if not row:
                continue
            non_empty = 0
            text_cells = 0
            for cell in row:
                if cell is not None and str(cell).strip():
                    non_empty += 1
                    cell_str = str(cell).strip()
                    # Headers are text; data rows have numbers
                    try:
                        float(cell_str.replace(",", "").replace("(", "").replace(")", ""))
                    except ValueError:
                        text_cells += 1
            # Score: text cells weighted more (headers are labels)
            score = text_cells * 2 + non_empty
            if score > best_score:
                best_score = score
                best_idx = i

        return best_idx

    @staticmethod
    def _row_to_dict(row: tuple, headers: List[str]) -> Dict[str, str]:
        """Convert a raw row tuple to a dict keyed by headers."""
        row_dict: Dict[str, str] = {}
        for i, val in enumerate(row):
            if i < len(headers):
                if val is None:
                    row_dict[headers[i]] = ""
                elif isinstance(val, datetime):
                    row_dict[headers[i]] = val.strftime("%Y-%m-%d")
                elif isinstance(val, date):
                    row_dict[headers[i]] = val.strftime("%Y-%m-%d")
                else:
                    row_dict[headers[i]] = str(val).strip()
        return row_dict

    @staticmethod
    def parse(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
        """
        Parse Excel file content and return (headers, rows).
        Uses smart header detection to handle D365 report-format exports
        that have title / metadata rows before the actual column headers.
        """
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return [], []

        # Read all rows into memory
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return [], []

        # Smart header detection: find the row that looks most like column headers
        header_idx = ExcelParser._detect_header_row(all_rows)
        header_row = all_rows[header_idx]

        headers = [
            str(h).strip() if h is not None else f"column_{i}"
            for i, h in enumerate(header_row)
        ]

        # Remove trailing auto-generated column names
        while headers and headers[-1].startswith("column_"):
            headers.pop()

        # If too few real headers detected, fall back to row 0
        real_headers = [h for h in headers if not h.startswith("column_")]
        if len(real_headers) < 2 and header_idx != 0:
            header_idx = 0
            header_row = all_rows[0]
            headers = [
                str(h).strip() if h is not None else f"column_{i}"
                for i, h in enumerate(header_row)
            ]
            while headers and headers[-1].startswith("column_"):
                headers.pop()

        # Parse data rows (everything after the header row)
        rows = []
        for row in all_rows[header_idx + 1:]:
            row_dict = ExcelParser._row_to_dict(row, headers)
            if any(v for v in row_dict.values()):  # Skip fully empty rows
                rows.append(row_dict)

        return headers, rows

    @staticmethod
    def parse_all_sheets(content: bytes) -> Dict[str, Tuple[List[str], List[Dict[str, str]]]]:
        """
        Parse all sheets from an Excel file.
        Returns {sheet_name: (headers, rows)}.
        """
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                continue

            headers = [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(header_row)]
            rows = []
            for row in rows_iter:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        if val is None:
                            row_dict[headers[i]] = ""
                        elif isinstance(val, datetime):
                            row_dict[headers[i]] = val.strftime("%Y-%m-%d")
                        elif isinstance(val, date):
                            row_dict[headers[i]] = val.strftime("%Y-%m-%d")
                        else:
                            row_dict[headers[i]] = str(val).strip()
                if any(v for v in row_dict.values()):
                    rows.append(row_dict)

            if rows:
                result[sheet_name] = (headers, rows)

        wb.close()
        return result


# =============================================
# CSV Parser
# =============================================

class CSVParser:
    """Parse and auto-detect CSV structure."""

    @staticmethod
    def parse(content: str, encoding: str = "utf-8") -> Tuple[List[str], List[Dict[str, str]]]:
        """
        Parse CSV content and return (headers, rows).
        Auto-detects delimiter (comma, semicolon, tab, pipe).
        """
        # Detect delimiter
        sample = content[:2000]
        delimiters = {",": 0, ";": 0, "\t": 0, "|": 0}
        for d in delimiters:
            delimiters[d] = sample.count(d)
        delimiter = max(delimiters, key=delimiters.get)
        if delimiters[delimiter] == 0:
            delimiter = ","

        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        headers = reader.fieldnames or []
        rows = []
        for row in reader:
            rows.append(dict(row))

        return headers, rows

    # Entity-specific alias overrides — applied BEFORE the global alias table.
    # For invoices/payments, "customer_code" must resolve to "customer_external_id"
    # rather than the entity's own "external_id".
    ENTITY_ALIASES = {
        "invoices": {
            "customer_code": "customer_external_id",
            "cust_code": "customer_external_id",
            "account_code": "customer_external_id",
            "customer_id": "customer_external_id",
            "cust_id": "customer_external_id",
            "code": "customer_external_id",
            # D365 specific
            "account": "customer_external_id",
            "customer_account": "customer_external_id",
            "order_account": "customer_external_id",
            "invoice_account": "customer_external_id",
            "customer": "customer_name",
            "delivery_name": "customer_name",
            "sales_order": "invoice_number",
            "order_number": "invoice_number",
            "sales_id": "invoice_number",
            "document_number": "invoice_number",
            "voucher": "invoice_number",
        },
        "payments": {
            "customer_code": "customer_external_id",
            "cust_code": "customer_external_id",
            "account_code": "customer_external_id",
            "customer_id": "customer_external_id",
            "cust_id": "customer_external_id",
            "code": "customer_external_id",
            # D365 specific
            "account": "customer_external_id",
            "customer_account": "customer_external_id",
            "customer": "customer_name",
        },
        "collections": {
            "customer_code": "customer_external_id",
            "cust_code": "customer_external_id",
            "customer_id": "customer_external_id",
            "cust_id": "customer_external_id",
            # D365 specific
            "account": "customer_external_id",
            "customer_account": "customer_external_id",
            "customer": "customer_name",
        },
        "disputes": {
            "customer_code": "customer_external_id",
            "cust_code": "customer_external_id",
            "customer_id": "customer_external_id",
            "cust_id": "customer_external_id",
            # D365 specific
            "account": "customer_external_id",
            "customer_account": "customer_external_id",
            "customer": "customer_name",
        },
        "credit_limits": {
            "customer_code": "customer_external_id",
            "cust_code": "customer_external_id",
            "customer_id": "customer_external_id",
            "cust_id": "customer_external_id",
            # D365 specific
            "account": "customer_external_id",
            "customer_account": "customer_external_id",
            "customer": "customer_name",
        },
    }

    @staticmethod
    def auto_map(csv_headers: List[str], entity_type: str) -> Dict[str, str]:
        """
        Auto-map CSV headers to entity fields using aliases.
        Returns {csv_header: entity_field} mapping.
        Only the first match wins per target field (no duplicate targets).
        Headers mapping to '_skip' are excluded from results.
        """
        fields = ENTITY_FIELDS.get(entity_type, {})
        entity_aliases = CSVParser.ENTITY_ALIASES.get(entity_type, {})
        mapping = {}
        used_targets: set = set()  # Prevent duplicate target assignments

        for header in csv_headers:
            normalized = header.lower().strip().replace(" ", "_").replace("-", "_")
            target = None

            # Direct match
            if normalized in fields:
                target = normalized

            # Entity-specific alias match (highest priority)
            if target is None and normalized in entity_aliases:
                alias_target = entity_aliases[normalized]
                if alias_target in fields:
                    target = alias_target

            # Global alias match
            if target is None and normalized in HEADER_ALIASES:
                alias_target = HEADER_ALIASES[normalized]
                if alias_target in fields:
                    target = alias_target

            # Fuzzy partial match (lowest priority)
            if target is None:
                for field_name in fields:
                    if field_name in normalized or normalized in field_name:
                        target = field_name
                        break

            # Skip internal markers and deduplicate targets
            if target and target != "_skip" and target not in used_targets:
                mapping[header] = target
                used_targets.add(target)

        return mapping


# =============================================
# Import Pipeline
# =============================================

DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"]


def _parse_date(value: str) -> Optional[date]:
    """Try multiple date formats."""
    if not value or not value.strip():
        return None
    value = value.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: str) -> Optional[Decimal]:
    if not value or not value.strip():
        return None
    value = value.strip().replace(",", "").replace(" ", "")
    # Handle parentheses for negatives: (500) -> -500
    if value.startswith("(") and value.endswith(")"):
        value = "-" + value[1:-1]
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def _parse_int(value: str) -> Optional[int]:
    if not value or not value.strip():
        return None
    try:
        return int(float(value.strip()))
    except (ValueError, TypeError):
        return None


class ImportPipeline:
    """
    Processes mapped CSV rows into database entities.
    Validates, transforms, and creates records with full error tracking.
    """

    SOURCE_SYSTEM = "CSV_IMPORT"

    async def preview(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        entity_type: str,
        rows: List[Dict[str, str]],
        mapping: Dict[str, str],
        max_preview: int = 10,
    ) -> Dict:
        """
        Validate and preview mapped data before import.
        Returns parsed rows, validation errors, and statistics.
        """
        fields = ENTITY_FIELDS.get(entity_type, {})
        required_fields = [k for k, v in fields.items() if v[2]]
        mapped_target_fields = set(mapping.values())

        # Check required fields coverage
        missing_required = [f for f in required_fields if f not in mapped_target_fields]

        preview_rows = []
        errors = []
        warnings = []

        for idx, row in enumerate(rows[:max_preview]):
            parsed = {}
            row_errors = []

            for csv_col, target_field in mapping.items():
                raw_value = row.get(csv_col, "").strip()
                if not raw_value:
                    continue

                field_def = fields.get(target_field)
                if not field_def:
                    continue

                _, field_type, is_required, _ = field_def

                if field_type == "date":
                    parsed_val = _parse_date(raw_value)
                    if parsed_val is None and raw_value:
                        row_errors.append(f"Row {idx + 1}: Cannot parse date '{raw_value}' for '{target_field}'")
                    else:
                        parsed[target_field] = str(parsed_val) if parsed_val else None
                elif field_type == "decimal":
                    parsed_val = _parse_decimal(raw_value)
                    if parsed_val is None and raw_value:
                        row_errors.append(f"Row {idx + 1}: Cannot parse number '{raw_value}' for '{target_field}'")
                    else:
                        parsed[target_field] = str(parsed_val) if parsed_val else None
                elif field_type == "int":
                    parsed_val = _parse_int(raw_value)
                    if parsed_val is None and raw_value:
                        row_errors.append(f"Row {idx + 1}: Cannot parse integer '{raw_value}' for '{target_field}'")
                    else:
                        parsed[target_field] = parsed_val
                else:
                    parsed[target_field] = raw_value

            # Check required fields in this row
            for req in required_fields:
                if req in mapped_target_fields and not parsed.get(req):
                    row_errors.append(f"Row {idx + 1}: Required field '{req}' is empty")

            preview_rows.append({"row_number": idx + 1, "parsed": parsed, "errors": row_errors})
            errors.extend(row_errors)

        return {
            "entity_type": entity_type,
            "total_rows": len(rows),
            "preview_rows": preview_rows,
            "missing_required_fields": missing_required,
            "mapped_fields": list(mapped_target_fields),
            "unmapped_headers": [h for h in rows[0].keys() if h not in mapping] if rows else [],
            "errors": errors,
            "warnings": warnings,
            "can_import": len(missing_required) == 0,
        }

    async def execute(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        user_id: UUID,
        entity_type: str,
        rows: List[Dict[str, str]],
        mapping: Dict[str, str],
    ) -> Dict:
        """Execute the import — create entities from CSV rows."""

        if entity_type == "customers":
            return await self._import_customers(db, tenant_id, user_id, rows, mapping)
        elif entity_type == "invoices":
            return await self._import_invoices(db, tenant_id, user_id, rows, mapping)
        elif entity_type == "payments":
            return await self._import_payments(db, tenant_id, user_id, rows, mapping)
        elif entity_type == "collections":
            return await self._import_collections(db, tenant_id, user_id, rows, mapping)
        elif entity_type == "disputes":
            return await self._import_disputes(db, tenant_id, user_id, rows, mapping)
        elif entity_type == "credit_limits":
            return await self._import_credit_limits(db, tenant_id, user_id, rows, mapping)
        else:
            return {"error": f"Unsupported entity type: {entity_type}"}

    async def _import_customers(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        updated = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "customers")
                if not data.get("name"):
                    errors.append(f"Row {idx + 1}: Missing required field 'name'")
                    skipped += 1
                    continue

                # Check for existing customer by external_id or name
                existing = None
                if data.get("external_id"):
                    result = await db.execute(
                        select(Customer).where(
                            Customer.tenant_id == tenant_id,
                            Customer.external_id == data["external_id"],
                            Customer.source_system == self.SOURCE_SYSTEM,
                        )
                    )
                    existing = result.scalar_one_or_none()

                if existing:
                    # Update existing
                    for field, value in data.items():
                        if value is not None and hasattr(existing, field):
                            setattr(existing, field, value)
                    existing.updated_by = user_id
                    updated += 1
                else:
                    # Create new
                    customer = Customer(
                        tenant_id=tenant_id,
                        created_by=user_id,
                        source_system=self.SOURCE_SYSTEM,
                        name=data["name"],
                        name_ar=data.get("name_ar"),
                        trade_name=data.get("trade_name"),
                        external_id=data.get("external_id"),
                        tax_id=data.get("tax_id"),
                        status=data.get("status", "active"),
                        industry=data.get("industry"),
                        segment=data.get("segment"),
                        territory=data.get("territory"),
                        region=data.get("region"),
                        country=data.get("country", "AE"),
                        city=data.get("city"),
                        address=data.get("address"),
                        phone=data.get("phone"),
                        email=data.get("email"),
                        website=data.get("website"),
                        currency=data.get("currency", "AED"),
                        payment_terms_days=data.get("payment_terms_days", 30),
                        credit_limit=data.get("credit_limit", Decimal("0")),
                        tags=["csv_import"],
                    )
                    db.add(customer)
                    created += 1

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id,
            user_id=user_id,
            action="CSV_IMPORT",
            entity_type="customers",
            after_state={"created": created, "updated": updated, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {
            "entity_type": "customers",
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    async def _import_invoices(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "invoices")

                # Resolve customer
                customer_id = await self._resolve_customer(
                    db, tenant_id,
                    data.get("customer_external_id"),
                    data.get("customer_name"),
                )
                if not customer_id:
                    errors.append(f"Row {idx + 1}: Cannot resolve customer (ext_id={data.get('customer_external_id')}, name={data.get('customer_name')})")
                    skipped += 1
                    continue

                if not data.get("invoice_number"):
                    errors.append(f"Row {idx + 1}: Missing invoice_number")
                    skipped += 1
                    continue

                inv_date = _parse_date(data.get("invoice_date", ""))
                due_dt = _parse_date(data.get("due_date", ""))
                amount = _parse_decimal(data.get("amount", ""))

                if not inv_date or not due_dt or not amount:
                    errors.append(f"Row {idx + 1}: Missing or invalid invoice_date, due_date, or amount")
                    skipped += 1
                    continue

                # Check duplicate
                dup = await db.execute(
                    select(Invoice).where(
                        Invoice.tenant_id == tenant_id,
                        Invoice.invoice_number == data["invoice_number"],
                    )
                )
                if dup.scalar_one_or_none():
                    errors.append(f"Row {idx + 1}: Invoice '{data['invoice_number']}' already exists")
                    skipped += 1
                    continue

                tax = _parse_decimal(data.get("tax_amount", "")) or Decimal("0")
                discount = _parse_decimal(data.get("discount_amount", "")) or Decimal("0")
                total = amount + tax - discount

                # Aging
                today = date.today()
                if due_dt >= today:
                    days_overdue, aging_bucket = 0, "current"
                else:
                    days_overdue = (today - due_dt).days
                    if days_overdue <= 30: aging_bucket = "1-30"
                    elif days_overdue <= 60: aging_bucket = "31-60"
                    elif days_overdue <= 90: aging_bucket = "61-90"
                    else: aging_bucket = "90+"

                status_str = data.get("status", "open").lower()
                inv_status = InvoiceStatus.OVERDUE if days_overdue > 0 and status_str == "open" else status_str

                invoice = Invoice(
                    tenant_id=tenant_id,
                    created_by=user_id,
                    customer_id=customer_id,
                    invoice_number=data["invoice_number"],
                    external_id=data.get("external_id"),
                    source_system=self.SOURCE_SYSTEM,
                    po_number=data.get("po_number"),
                    invoice_date=inv_date,
                    due_date=due_dt,
                    currency=data.get("currency", "AED"),
                    amount=total,
                    tax_amount=tax,
                    discount_amount=discount,
                    amount_paid=Decimal("0"),
                    amount_remaining=total,
                    status=inv_status,
                    days_overdue=days_overdue,
                    aging_bucket=aging_bucket,
                    notes=data.get("notes"),
                )
                db.add(invoice)
                created += 1

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id, user_id=user_id,
            action="CSV_IMPORT", entity_type="invoices",
            after_state={"created": created, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {"entity_type": "invoices", "total_rows": len(rows), "created": created, "skipped": skipped, "errors": errors}

    async def _import_payments(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        matched = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "payments")

                customer_id = await self._resolve_customer(
                    db, tenant_id,
                    data.get("customer_external_id"),
                    data.get("customer_name"),
                )
                if not customer_id:
                    errors.append(f"Row {idx + 1}: Cannot resolve customer")
                    skipped += 1
                    continue

                pay_date = _parse_date(data.get("payment_date", ""))
                amount = _parse_decimal(data.get("amount", ""))
                if not pay_date or not amount:
                    errors.append(f"Row {idx + 1}: Missing payment_date or amount")
                    skipped += 1
                    continue

                # Try to match invoice
                invoice_id = None
                is_matched = False
                match_confidence = None

                if data.get("invoice_number"):
                    inv_result = await db.execute(
                        select(Invoice).where(
                            Invoice.tenant_id == tenant_id,
                            Invoice.invoice_number == data["invoice_number"],
                        )
                    )
                    inv = inv_result.scalar_one_or_none()
                    if inv:
                        invoice_id = inv.id
                        is_matched = True
                        match_confidence = 1.0

                        # Update invoice
                        inv.amount_paid = (inv.amount_paid or Decimal("0")) + amount
                        inv.amount_remaining = inv.amount - inv.amount_paid
                        if inv.amount_remaining <= 0:
                            inv.amount_remaining = Decimal("0")
                            inv.status = InvoiceStatus.PAID
                        else:
                            inv.status = InvoiceStatus.PARTIALLY_PAID
                        matched += 1

                # Resolve payment method
                method = data.get("payment_method")
                try:
                    method = PaymentMethod(method) if method else None
                except ValueError:
                    method = None

                payment = Payment(
                    tenant_id=tenant_id,
                    created_by=user_id,
                    customer_id=customer_id,
                    invoice_id=invoice_id,
                    external_id=data.get("external_id"),
                    source_system=self.SOURCE_SYSTEM,
                    payment_date=pay_date,
                    amount=amount,
                    currency=data.get("currency", "AED"),
                    payment_method=method,
                    reference_number=data.get("reference_number"),
                    bank_reference=data.get("bank_reference"),
                    is_matched=is_matched,
                    matched_at=datetime.now(timezone.utc).isoformat() if is_matched else None,
                    match_confidence=match_confidence,
                    notes=data.get("notes"),
                )
                db.add(payment)
                created += 1

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id, user_id=user_id,
            action="CSV_IMPORT", entity_type="payments",
            after_state={"created": created, "matched": matched, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {"entity_type": "payments", "total_rows": len(rows), "created": created, "matched": matched, "skipped": skipped, "errors": errors}

    async def _import_collections(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "collections")

                customer_id = await self._resolve_customer(
                    db, tenant_id,
                    data.get("customer_external_id"),
                    data.get("customer_name"),
                )
                if not customer_id:
                    errors.append(f"Row {idx + 1}: Cannot resolve customer")
                    skipped += 1
                    continue

                action_date = _parse_date(data.get("action_date", ""))
                if not action_date:
                    errors.append(f"Row {idx + 1}: Missing or invalid action_date")
                    skipped += 1
                    continue

                # Resolve action type
                action_type_str = (data.get("action_type") or "phone_call").lower().strip()
                try:
                    action_type = CollectionAction(action_type_str)
                except ValueError:
                    action_type = CollectionAction.PHONE_CALL

                # Resolve invoice if provided
                invoice_id = None
                if data.get("invoice_number"):
                    inv_result = await db.execute(
                        select(Invoice).where(
                            Invoice.tenant_id == tenant_id,
                            Invoice.invoice_number == data["invoice_number"],
                        )
                    )
                    inv = inv_result.scalar_one_or_none()
                    if inv:
                        invoice_id = inv.id

                ptp_date = _parse_date(data.get("ptp_date", ""))
                ptp_amount = _parse_decimal(data.get("ptp_amount", ""))

                activity = CollectionActivity(
                    tenant_id=tenant_id,
                    created_by=user_id,
                    customer_id=customer_id,
                    invoice_id=invoice_id,
                    collector_id=user_id,
                    action_type=action_type,
                    action_date=action_date,
                    notes=data.get("notes"),
                    ptp_date=ptp_date,
                    ptp_amount=ptp_amount,
                )
                db.add(activity)
                created += 1

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id, user_id=user_id,
            action="CSV_IMPORT", entity_type="collections",
            after_state={"created": created, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {"entity_type": "collections", "total_rows": len(rows), "created": created, "skipped": skipped, "errors": errors}

    async def _import_disputes(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "disputes")

                customer_id = await self._resolve_customer(
                    db, tenant_id,
                    data.get("customer_external_id"),
                    data.get("customer_name"),
                )
                if not customer_id:
                    errors.append(f"Row {idx + 1}: Cannot resolve customer")
                    skipped += 1
                    continue

                dispute_number = data.get("dispute_number")
                if not dispute_number:
                    errors.append(f"Row {idx + 1}: Missing dispute_number")
                    skipped += 1
                    continue

                amount = _parse_decimal(data.get("amount", ""))
                if not amount:
                    errors.append(f"Row {idx + 1}: Missing or invalid amount")
                    skipped += 1
                    continue

                # Resolve reason
                reason_str = (data.get("reason") or "other").lower().strip()
                try:
                    reason = DisputeReason(reason_str)
                except ValueError:
                    reason = DisputeReason.OTHER

                # Resolve status
                status_str = (data.get("status") or "open").lower().strip()
                try:
                    dispute_status = DisputeStatus(status_str)
                except ValueError:
                    dispute_status = DisputeStatus.OPEN

                # Resolve invoice if provided
                invoice_id = None
                if data.get("invoice_number"):
                    inv_result = await db.execute(
                        select(Invoice).where(
                            Invoice.tenant_id == tenant_id,
                            Invoice.invoice_number == data["invoice_number"],
                        )
                    )
                    inv = inv_result.scalar_one_or_none()
                    if inv:
                        invoice_id = inv.id

                sla_date = _parse_date(data.get("sla_due_date", ""))

                dispute = Dispute(
                    tenant_id=tenant_id,
                    created_by=user_id,
                    customer_id=customer_id,
                    invoice_id=invoice_id,
                    dispute_number=dispute_number,
                    reason=reason,
                    reason_detail=data.get("reason_detail"),
                    status=dispute_status,
                    amount=amount,
                    currency=data.get("currency", "AED"),
                    priority=data.get("priority", "medium"),
                    sla_due_date=sla_date,
                )
                db.add(dispute)
                created += 1

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id, user_id=user_id,
            action="CSV_IMPORT", entity_type="disputes",
            after_state={"created": created, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {"entity_type": "disputes", "total_rows": len(rows), "created": created, "skipped": skipped, "errors": errors}

    async def _import_credit_limits(
        self, db: AsyncSession, tenant_id: UUID, user_id: UUID,
        rows: List[Dict[str, str]], mapping: Dict[str, str],
    ) -> Dict:
        created = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows):
            try:
                data = self._extract_mapped_values(row, mapping, "credit_limits")

                customer_id = await self._resolve_customer(
                    db, tenant_id,
                    data.get("customer_external_id"),
                    data.get("customer_name"),
                )
                if not customer_id:
                    errors.append(f"Row {idx + 1}: Cannot resolve customer")
                    skipped += 1
                    continue

                requested_limit = _parse_decimal(data.get("requested_limit", ""))
                if not requested_limit:
                    errors.append(f"Row {idx + 1}: Missing or invalid requested_limit")
                    skipped += 1
                    continue

                # Look up current credit limit from customer
                cust_result = await db.execute(
                    select(Customer.credit_limit).where(Customer.id == customer_id)
                )
                current_limit = cust_result.scalar_one_or_none() or Decimal("0")

                # Resolve approval status
                approval_str = (data.get("approval_status") or "pending").lower().strip()
                try:
                    approval_status = CreditApprovalStatus(approval_str)
                except ValueError:
                    approval_status = CreditApprovalStatus.PENDING

                request = CreditLimitRequest(
                    tenant_id=tenant_id,
                    created_by=user_id,
                    customer_id=customer_id,
                    requested_by_id=user_id,
                    current_limit=current_limit,
                    requested_limit=requested_limit,
                    currency=data.get("currency", "AED"),
                    justification=data.get("justification"),
                    approval_status=approval_status,
                )
                db.add(request)
                created += 1

                # If pre-approved, also update customer limit
                if approval_status == CreditApprovalStatus.APPROVED:
                    cust_update = await db.execute(
                        select(Customer).where(Customer.id == customer_id)
                    )
                    customer = cust_update.scalar_one_or_none()
                    if customer:
                        customer.credit_limit = requested_limit
                        request.approved_limit = requested_limit
                        request.approved_by_id = user_id
                        request.approved_at = datetime.now(timezone.utc).isoformat()

            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
                skipped += 1

        audit = AuditLog(
            tenant_id=tenant_id, user_id=user_id,
            action="CSV_IMPORT", entity_type="credit_limits",
            after_state={"created": created, "skipped": skipped, "errors": len(errors)},
        )
        db.add(audit)
        await db.commit()

        return {"entity_type": "credit_limits", "total_rows": len(rows), "created": created, "skipped": skipped, "errors": errors}

    def _extract_mapped_values(
        self, row: Dict[str, str], mapping: Dict[str, str], entity_type: str
    ) -> Dict[str, Any]:
        """Extract and type-cast values from a CSV row using the mapping."""
        fields = ENTITY_FIELDS.get(entity_type, {})
        result = {}

        for csv_col, target_field in mapping.items():
            raw = row.get(csv_col, "").strip()
            if not raw:
                continue

            field_def = fields.get(target_field)
            if not field_def:
                result[target_field] = raw
                continue

            _, field_type, _, _ = field_def

            if field_type == "decimal":
                result[target_field] = str(_parse_decimal(raw)) if _parse_decimal(raw) is not None else raw
            elif field_type == "int":
                result[target_field] = _parse_int(raw)
            elif field_type == "date":
                result[target_field] = str(raw)  # Parse later for flexibility
            else:
                result[target_field] = raw

        return result

    async def _resolve_customer(
        self, db: AsyncSession, tenant_id: UUID,
        external_id: Optional[str], name: Optional[str],
    ) -> Optional[UUID]:
        """Resolve customer_id from external_id or name."""
        if external_id:
            result = await db.execute(
                select(Customer.id).where(
                    Customer.tenant_id == tenant_id,
                    Customer.external_id == external_id,
                )
            )
            row = result.scalar_one_or_none()
            if row:
                return row

        if name:
            result = await db.execute(
                select(Customer.id).where(
                    Customer.tenant_id == tenant_id,
                    Customer.name == name,
                )
            )
            row = result.scalar_one_or_none()
            if row:
                return row

        return None


# Singletons
csv_parser = CSVParser()
excel_parser = ExcelParser()
import_pipeline = ImportPipeline()
